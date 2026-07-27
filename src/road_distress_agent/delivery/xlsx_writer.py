"""Render a CostSheet to an auditable xlsx workbook (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from road_distress_agent.delivery.costing import CostSheet

_BOLD = Font(bold=True)


def write_cost_workbook(
    sheet: CostSheet, *, project: dict[str, Any], path: Path, locale: str = "zh-CN"
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = _label("sheet_title", locale)

    _write_title(ws, project, locale)
    _write_defect_table(ws, sheet, locale)
    _write_mobilization_table(ws, sheet, locale)
    _write_totals(ws, sheet, locale)

    wb.save(path)
    return path


def _write_title(ws: Any, project: dict[str, Any], locale: str) -> None:
    ws.append([f"{_label('title', locale)}: {project.get('name', '')}"])
    ws["A1"].font = _BOLD
    ws.append(
        [
            _label("segment", locale),
            project.get("segment") or "—",
            _label("crew", locale),
            project.get("crew") or "—",
        ]
    )
    ws.append([])


def _write_defect_table(ws: Any, sheet: CostSheet, locale: str) -> None:
    _append_header(ws, _headers("defect", locale))
    for line in sheet.defect_lines:
        ws.append(
            [
                _dialogue_id(line),
                line.defect_category,
                line.norm_code,
                line.process_name,
                line.calculation_unit,
                line.quantity,
                line.unit_cost_cny,
                line.subtotal_cny,
            ]
        )
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            _label("defect_subtotal", locale),
            sheet.defect_subtotal_cny,
        ]
    )
    ws.append([])


def _write_mobilization_table(ws: Any, sheet: CostSheet, locale: str) -> None:
    ws.append([_label("mobilization_title", locale)])
    ws.cell(row=ws.max_row, column=1).font = _BOLD
    _append_header(ws, _headers("mobilization", locale))
    for line in sheet.mobilization_lines:
        ws.append([line.rule_code, line.cost_name, line.amount_cny, line.aggregation_rule])
    ws.append(["", _label("mobilization_subtotal", locale), sheet.mobilization_subtotal_cny, ""])
    ws.append([])


def _write_totals(ws: Any, sheet: CostSheet, locale: str) -> None:
    ws.append([_label("total", locale), sheet.total_cny])
    ws.cell(row=ws.max_row, column=1).font = _BOLD


def _append_header(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = _BOLD


def _dialogue_id(line: Any) -> str:
    dialogue_id = getattr(line, "source_thread_id", None)
    if not dialogue_id:
        raise ValueError(f"cost line {line.record_id} missing source_thread_id.")
    return str(dialogue_id)


def _headers(kind: str, locale: str) -> list[str]:
    labels = _LABELS_EN if locale == "en-US" else _LABELS_ZH
    return list(labels[f"{kind}_headers"])


def _label(key: str, locale: str) -> str:
    labels = _LABELS_EN if locale == "en-US" else _LABELS_ZH
    return str(labels[key])


_LABELS_ZH = {
    "sheet_title": "工程量与造价",
    "title": "巡查任务造价表",
    "segment": "路段桩号",
    "crew": "责任班组",
    "defect_subtotal": "病害小计(元)",
    "mobilization_title": "进场费(路段级聚合，每任务只计一次)",
    "mobilization_subtotal": "进场费小计(元)",
    "total": "合计(元)",
    "defect_headers": [
        "对话ID",
        "病害类别",
        "定额编号",
        "工艺",
        "计量单位",
        "工程量",
        "单价(元)",
        "小计(元)",
    ],
    "mobilization_headers": ["规则编号", "进场费名称", "金额(元)", "聚合规则"],
}
_LABELS_EN = {
    "sheet_title": "Quantity and Cost",
    "title": "Inspection Cost Sheet",
    "segment": "Road Segment / Stake",
    "crew": "Responsible Crew",
    "defect_subtotal": "Defect Subtotal (CNY)",
    "mobilization_title": "Mobilization (segment-level, counted once per task)",
    "mobilization_subtotal": "Mobilization Subtotal (CNY)",
    "total": "Total (CNY)",
    "defect_headers": [
        "Dialogue ID",
        "Distress Type",
        "Norm Code",
        "Process",
        "Unit",
        "Quantity",
        "Unit Cost (CNY)",
        "Subtotal (CNY)",
    ],
    "mobilization_headers": ["Rule Code", "Mobilization Item", "Amount (CNY)", "Aggregation Rule"],
}
